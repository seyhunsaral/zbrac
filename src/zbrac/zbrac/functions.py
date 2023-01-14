#    This file is part of zBrac.
#
#    zBrac is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    zBrac is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with zBrac.  If not, see <https://www.gnu.org/licenses/>.

import sys
import os
import re
import xlsxwriter
import openpyxl
import sys
import os


def load_treatment_file(filepath, encoding = "UTF-8"):
    try:
        with open(filepath, encoding=encoding) as file: # Turkish
            filedata = file.read()
        file.close()
    except:
        print('! Warning: Non readable charachters exist for encoding ' + encoding + '. \n Ignoring non convertible chars.\n Please check if you have selected right encoding')
        with open(filepath, encoding=encoding, errors='ignore') as file: # Turkish
            filedata = file.read()
        file.close()
    print('Loaded treatment file:' + filepath)
    print('')
    return(filedata)


def save_treatment_file(treatment_text, filepath, encoding = "UTF-8"):
    try:
        with open(filepath, 'w', encoding=encoding) as file:
            file.write(treatment_text)
        file.close()
    except:
        print('! Warning: Non writable characters exist for selected encoding ' + encoding + '. \n Ignoring non convertible chars.\n Please check if you have selected right encoding')
        with open(filepath, 'w', encoding=encoding, errors='ignore') as file:
            file.write(treatment_text)
        file.close()
    print('Saved treatment file:' + filepath)


def remove_escapes(stringvar):
    return(stringvar.replace('\\"','\"').replace('\\\\','\\'))

def add_escapes(stringvar):
    return(stringvar.replace('\\','\\\\').replace('\"','\\"'))

def sanitize_r_n(stringvar):
    return(stringvar.replace('\\\\r\\\\n','\\r\\n'))

def add_escapes_dict(dictionary):
    ## while things like \cf \par are quoted in the treatment file, for some reason, \r and \n dont. 
    ## before 1.0.8 version, it was missing files with \r\n
    escaped_dict = dict()

    for key in dictionary:
        escaped_dict[sanitize_r_n(add_escapes(key))] = sanitize_r_n(add_escapes(dictionary[key]))
    
    return escaped_dict


def get_matched_entries(textblock):
    matched_items = re.findall("\[\[.*?\]\]", textblock)
    if not matched_items:
        return(False)
    unique_matched_items = list(set(matched_items))
    unique_matched_items.sort()
    print('    Grabbed following keywords:')
    print('  ' + '-' * 35)
    for item in unique_matched_items: print('    ' + remove_escapes(item))
    print('-' * 35)
    print(str(len(unique_matched_items)) + ' items in total')
    print('')
    return(unique_matched_items)


def create_own_list(keywordlist):
    try:
        escaped_keywordlist = list(map(remove_escapes, keywordlist))
        stripped_keywordlist = list(map(lambda each:each.replace('[[','').replace(']]',''), escaped_keywordlist))
        return([escaped_keywordlist,stripped_keywordlist])
    except:
        if keywordlist == False:
            print('!!Error: Some problem occurred. This might be due to an empty list of keywords.')
        return(keywordlist)


def list_to_dict(keywordlist):
    return(dict(zip(keywordlist[0],keywordlist[1])))


def list_to_xlsx(keywordlist, savepath):
    try:
        workbook = xlsxwriter.Workbook(savepath)
        worksheet = workbook.add_worksheet()
        rowno = 0
        for col1, col2 in zip(keywordlist[0], keywordlist[1]):
            worksheet.write(rowno, 0, col1)
            worksheet.write(rowno, 1, col2)
            rowno += 1
        workbook.close()
        print('Succesfully saved to ' + savepath)
        print('')
    except:
        print('!! Error: A problem occured while saving the file')


def xlsx_to_dictionary(filepath):
    try:
        pattern = re.compile("\[\[.*?\]\]")

        language_dict = dict()
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        for row in worksheet.iter_rows():
#            print("hello", row[0].value)
            if row[0].value is not None and row[1].value is not None:
                if (len(row) > 1):
                    if pattern.match(row[0].value):
                        language_dict[row[0].value] = str(row[1].value)

        workbook.close()
        print('Loaded language file:' + filepath)
        print('')
        print('    Grabbed following translations:')
        print('  ' + '-' * 35)
        for item in language_dict: print(" \n     " + item + "\n    -> " + language_dict[item])
        if (not len(language_dict) > 0):
           print("No keywords exist in the dictionary file")
           return
        print('  ' + '-' * 35)
        print(str(len(language_dict)) + ' items in total')
        
        language_dict = add_escapes_dict(language_dict)
        return (language_dict)
    except Exception as ex:
    #    print(ex)
        print("Error: A problem occured with reading excel file. Does it contain the keys as the first column and the text in second column?")
        return


def replace_from_dictionary(language_dict, sourcefile):
    new_sourcefile = sourcefile
    for key in language_dict:
        new_sourcefile = new_sourcefile.replace(key, language_dict[key])
        
    return(new_sourcefile)
        

def implement_language_file(path_treatment_in, path_language_in, path_treatment_out, strip_unmatched,encoding_input = 'UTF-8',encoding_output = 'UTF-8'):
    source_text = load_treatment_file(path_treatment_in, encoding_input)

    print(' ' + '======== Treatment file ========')
    matched_entries = get_matched_entries(source_text)
    if not matched_entries:
        print('! Error: No matched entries in treatment file. No file is generated')
        return(False)
        
    print('')
    print(' ' + '======== Language file ========')
    language_dict = xlsx_to_dictionary(path_language_in)
    if (not language_dict):
        print('Error with the language file. No treatment file has been generated.')
        return
    set_treatment = set(matched_entries)
    set_language = set(language_dict.keys())
    non_replaced_keys = set_treatment.difference(set_language)
    non_used_keys = set_language.difference(set_treatment)
    
    if non_replaced_keys:
        print('\n')
        print('! Non replaced keys in the treatment file:')
        for item in non_replaced_keys: print(' ' + item,end =",")
        print('')
        print('-' * 10)
        
    if non_used_keys:
        print('\n')
        print('!  Non used keys:')
        for item in non_used_keys: print(' ' + item,end =",")
        print('')
        print('-' * 10)

    if non_used_keys:
        print('! Warning: ' + str(len(non_used_keys)) + ' dictionary items were not used')
    
    if non_replaced_keys:
        print('')
        print('! Warning: ' + str(len(non_replaced_keys)) + ' keys were not replaced by a dictionary input')
        if strip_unmatched:
            print('        stripping brackets from non-replaced keys')
    
    target_text = replace_from_dictionary(language_dict, source_text)
    if strip_unmatched:
        target_text = strip_brackets(target_text)
        print('stipped brackets non-replaced keys succesfully') 
    
    save_treatment_file(target_text, path_treatment_out,encoding_output)

def strip_brackets(source_text):
    matched_entries = get_matched_entries(source_text)
    language_list = create_own_list(matched_entries)
    language_dict = add_escapes_dict(list_to_dict(language_list))
    target_text = replace_from_dictionary(language_dict, source_text)
    return(target_text)

def strip_brackets_file(path_treatment_in, path_treatment_out,encoding='UTF-8'):
    source_text = load_treatment_file(path_treatment_in,encoding)
    target_text = strip_brackets(source_text)
    save_treatment_file(target_text, path_treatment_out, encoding)

    
### Function below will be only used for future implementation of the command line functionality
# def generate_language_file(path_treatment_in, path_language_out):
    # source_text = load_treatment_file(path_treatment_in)
    # matched_entries = get_matched_entries(source_text)
    # if not matched_entries:
        # print('no file has been generated')
        # return(False)
    # language_list = create_own_list(matched_entries)
    # list_to_csv(language_list, path_language_out)
